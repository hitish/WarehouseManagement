import API.amazonscrap as amazon
import asyncio
import openpyxl
import pandas as pd
import time
import pdb

async def main(sheetName):
    workbook = openpyxl.load_workbook(sheetName)
    sheets = workbook.sheetnames
    records = []
    current_time = time.time_ns()
    for sheet in sheets:
        workSheet = workbook[sheet]
        row = workSheet.max_row
        colum = workSheet.max_column
        itemCount = 0
        for i in range(2, row-52):
            code = workSheet.cell(row=i, column=1).value
            if not code:
                break;
            print("Processing file {} row {}".format(sheetName, i))
            link = 'https://www.amazon.in/dp/{}'.format(code)
            result = await amazon.getProductDetails(link)
            quantityCount = int(workSheet.cell(row=i, column=4).value)
            for count in range(1, quantityCount+1):
                try:
                    record = {}
                    itemCount=itemCount+1
                    record["name"] = result["name"] if 'name' in result and result['name']!='' else workSheet.cell(row=i, column=2).value
                    record["id"] = code
                    record["productCode"] = '{}_{}_{}_{}'.format(code,hash(sheetName),current_time,itemCount)
                    record["subcategory"] = workSheet.cell(row=i, column=3).value
                    record["osp"] =result["current_price"]
                    record["mrp"] =result["original_price"]
                    record["ratings"] =result["rating"]
                    record["size"] =result["size"]
                    record["color"] =result["color"]
                    record["vendor"] ='amazon'
                    record["brand"] =result["brand"]
                    record["model"] =result["model"]
                #   record["sellerName"] = workSheet.cell(row=i, column=10).value
                #   record["sheetPrice"] = workSheet.cell(row=i, column=8).value
                    record["metadata"] =result["metadata"]
                    record["time"] = current_time
                    records.append(record)
                except Exception as e:
                    print(result)
                    error = 'Some error occured while fetching product detail in amazon. Error for page {} is {}'.format(link, e)
                    print(error)
                    
    print(records)
    df = pd.DataFrame.from_dict(records)
    df.to_csv('treadmill1.csv', index=False, encoding='utf-8')



asyncio.run(main('TREADMILL82.xlsx'))