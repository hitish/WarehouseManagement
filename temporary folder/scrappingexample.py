from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

driver = webdriver.Chrome(r"C:\programming\chromedriver_win32\chromedriver.exe") #Set the path to chromedriver

online_code=[]
products=[] #List to store name of the product
prices=[] #List to store price of the product
ratings=[] #List to store rating of the product
review=[]
cat=[]
metadata={}
productdetail={}    
category=""
data = openpyxl.load_workbook(r'kitchen_nagaur.xlsx') 
sh = data.active
product_name=""
print(sh.max_row)

#for i in range(2,sh.max_row-5):
for i in range(2,311):
    Ocode = sh.cell(row=i, column=1).value
    if not Ocode:
        Ocode = ""
        
    link = "https://amazon.in/dp/" + Ocode
    driver.get(link)
    content = driver.page_source
    soup = BeautifulSoup(content)
   

    try:
        product_name=soup.find('span', attrs={'id':'productTitle'}).text.strip().replace(';','')
        price_block=soup.find('div', attrs={'id':'corePriceDisplay_desktop_feature_div'})
        osp = price_block.find('span', attrs={'class':'a-price-whole'})
        osp = osp.text.strip()
    except AttributeError:
        osp = ""
    
    try:
        temp = soup.find('span', attrs={'id':'acrPopover'})
        review_text = temp.find('span', attrs={'class':'a-icon-alt'}).text.strip()
    except Exception as e:
             error = 'Error for fetching review for page is {}'.format(e)
             currentPrice = 0
             print(error)
   
    try:
        additionaldatatable = soup.find('table', attrs={'id':'productDetails_detailBullets_sections1'})
        for row in additionaldatatable.find_all('tr'):
            try:
                name = row.find(attrs={'class': 'prodDetSectionEntry'}).text.strip()
                value = row.find(attrs={'class': 'prodDetAttrValue'}).text.strip().replace('\u200e','')
                metadata[name]=value
            except:
                continue
    except Exception as e:
        error = 'Error for fetching additionaldata for page is {}'.format(e)
        print(error)
             
    try:
        productdetailtable = soup.find('div', attrs={'id':'detailBullets_feature_div'})
        for row in productdetailtable.find_all('li'):
            try:
                name = row.find('span',attrs={'class': 'a-text-bold'}).text.strip().replace('\n                                    \u200f\n                                        :\n                                    \u200e','')
                value = row.find('span',attrs={'class': ''}).text.strip()
                productdetail[name]=value
            except:
                continue
    except Exception as e:
        error = 'Error for fetching productdetail for page is {}'.format(e)
        print(error)
       

    if not metadata:
        metadata = productdetail
             
    try:
        category = metadata["Generic Name"].replace(';','')
    except Exception as e:
        error = 'Error for fetching additionaldata for page is {}'.format(e)
        print(error)
        
        
    online_code.append(Ocode)
    products.append(product_name)
    prices.append(osp)
    review.append(review_text[0:3])
    cat.append(category)
    
    
driver.close()   
df = pd.DataFrame({'Online Code':online_code,'Product Name':products,'review':review,'OSP':prices,'category':cat}) 
#print(df)
df.to_csv('kitchen_nagaur1.csv', index=False, encoding='utf-8')