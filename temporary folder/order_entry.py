import API.amazonscrap as amazon
import API.database_functions as dbf
import asyncio
import openpyxl
import pandas as pd
import time
import pdb
import mysql.connector as mysql
import json

async def main(sheetName):
    try:
       connection = await dbf.create_database_connection()
       cursor = connection.cursor()

    except Exception as e:
        error = 'not connected due to {}'.format(e)
        print(error)
        
    workbook = openpyxl.load_workbook(sheetName)
    sheets = workbook.sheetnames
    records = []
    current_time = time.time_ns()
    lot_id = 5  
    
    for sheet in sheets:
        workSheet = workbook[sheet]
        row = workSheet.max_row
        colum = workSheet.max_column
        itemCount = 0
        for i in range(2, 5):
            try:
                box_id = workSheet.cell(row=i, column=1).value
                online_code = workSheet.cell(row=i, column=2).value
                
                product_link = "https://amazon.in/dp/"+online_code
                result = await amazon.getProductDetails(product_link)
            
                product_detail_query = "select * from sd_product_details where product_id = '"+ online_code+"'"
                cursor.execute(product_detail_query)
                product_detail = cursor.fetchall()
                #print(product_detail)
                if not product_detail:
                
                    product_name = workSheet.cell(row=i, column=3).value
                    if not product_name:
                        product_name = result["name"]
                        
                    category = workSheet.cell(row=i, column=4).value
                    if not category:
                        category = result["category"]
                        
                    brand = workSheet.cell(row=i, column=8).value
                    if not category:
                        brand = result["brand"]
                    
                    rating = result["rating"]
                    
                    mrp = workSheet.cell(row=i, column=7).value
                    if not category:
                        category = result["original_price"]
                        
                    model = result["model"]
                    color = result["color"]
                    size = result["size"]
                    metadata = result["metadata"]
                    #product_entry_query = "INSERT INTO `sd_product_details` (`product_id`, `product_name`, `category_id`, `brand`, `rating`, `mrp`, `model`, `color`, `size`, `metadata`) VALUES ('"+ online_code +"', '"+ product_name + "', '1', '"+ brand + "', '"+ rating + "', '"+ mrp + "', '"+ model + "', '"+ color + "', '"+ size + "', '"+ metadata + "')"
                    product_entry_query = "INSERT INTO `sd_product_details` (`product_id`, `product_name`, `category_id`, `brand`, `rating`, `mrp`, `model`, `color`, `size`, `metadata`) VALUES ('{}', '{}', '1', '{}', '{}', '{}', '{}', '{}', '{}','{}')".format(online_code,product_name,brand,rating,mrp,model,color,size,str(metadata).replace("\'","\""))
                    
                    
                    try:
                        cursor.execute(product_entry_query)
                        response = connection.commit()
                    except Exception as e:
                        connection.rollback()
                    
                quantity = workSheet.cell(row=i, column=5).value
                sosp = workSheet.cell(row=i, column=6).value
                cosp = result["current_price"]
                print("product_detail_query")
                #order_entry_query = "INSERT INTO `sd_unchecked_stock` (`order_lot_no`, `box_id`, `online_code`, `quantity`, `sosp`, `cosp`) VALUES ( '"+ lot_id +"', '"+ box_id +"', '"+ online_code +"', '"+ quantity +"', '"+ sosp +"', '"+ cosp +"')"
                order_entry_query = "INSERT INTO `sd_unchecked_stock` (`order_lot_no`, `box_id`, `online_code`, `quantity`, `sosp`, `cosp`) VALUES ( '{}', '{}', '{}', '{}', '{}', '{}')".format(lot_id,box_id,online_code,quantity,sosp,cosp)
                
                try:
                    cursor.execute(order_entry_query)
                    response = connection.commit()
                except Exception as e:
                    print("order with code {} and quantity {} not added d")
                    connection.rollback()
                    
              
            except Exception as e:
                #print(result)
                error = 'Some error occured while fetching product detail in amazon. Error for page is {}'.format( e)
                print(error)
                    
                    
                    
    connection.close()
    #print(records)
    #df = pd.DataFrame.from_dict(records)
    #df.to_csv('treadmill1.csv', index=False, encoding='utf-8')



asyncio.run(main('lot_10.xlsx'))