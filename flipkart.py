import API.flipkartfinal as flipkart
import asyncio
import openpyxl
import pandas as pd
import time
import pdb

async def main(sheetName):
    workbook = openpyxl.load_workbook(sheetName)
    sheets = workbook.sheetnames
    records = []
    current_time = time.time_ns()
    for sheet in sheets:
        workSheet = workbook[sheet]
        row = workSheet.max_row
        colum = workSheet.max_column
        itemCount = 0
        for i in range(2, row+1):
            code = workSheet.cell(row=i, column=1).value
            if not code:
                break;
            print("Processing file {} row {}".format(sheetName, i))
            link = 'https://www.flipkart.com/product/p/item?pid={}&marketplace=FLIPKART&sattr[]=color&sattr[]=size&st=size'.format(code)
            result = await flipkart.getProductDetails(link)
            quantityCount = int(workSheet.cell(row=i, column=6).value)
            for count in range(1, quantityCount+1):
                try:
                    record = {}
                    itemCount=itemCount+1
                    record["name"] = result["name"] if 'name' in result and result['name']!='' else workSheet.cell(row=i, column=7).value
                    record["id"] = code
                    record["productCode"] = '{}_{}_{}_{}'.format(code,hash(sheetName),current_time,itemCount)
                    record["subcategory"] = workSheet.cell(row=i, column=11).value
                    record["link"] =link
                    record["osp"] =result["current_price"]
                    record["mrp"] =result["original_price"]
                    record["ratings"] =result["rating"]
                    record["discount"] =result["discount"]
                    record["size"] =result["size"]
                    record["color"] =result["color"]
                    record["vendor"] ='flipkart'
                    record["brand"] =result["brand"] if 'brand' in result and result['brand']!='' else workSheet.cell(row=i, column=9).value
                    record["sellerName"] = workSheet.cell(row=i, column=10).value
                    record["sheetPrice"] = workSheet.cell(row=i, column=8).value
                    record["metadata"] =result["metadata"]
                    record["time"] = current_time
                    records.append(record)
                except Exception as e:
                    print(result)
                    error = 'Some error occured while fetching product detail.Error for page {} is {}'.format(link, e)
                    print(error)
    df = pd.DataFrame.from_dict(records)
    df.to_csv('final_details.csv', index=False, encoding='utf-8')



asyncio.run(main('sports sankpa 16062022.xlsx'))
